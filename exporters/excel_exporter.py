import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import io

def export_to_excel(extracted_data, filename="financial_data.xlsx"):
    """
    Export extracted financial data to Excel file.
    
    Args:
        extracted_data: Dictionary containing financial data
        filename: Output filename
    
    Returns:
        BytesIO object for download or True if saved successfully
    """
    try:
        # Create a new workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create sheets for each financial statement
        sheets_data = {
            'Balance Sheet': extracted_data.get('balance_sheet', []),
            'Profit & Loss': extracted_data.get('profit_loss', []),
            'Cash Flow': extracted_data.get('cash_flow', []),
            'Notes': extracted_data.get('notes', [])
        }
        
        for sheet_name, data in sheets_data.items():
            ws = wb.create_sheet(title=sheet_name)
            
            if sheet_name == 'Notes':
                # Handle notes as text
                for idx, note in enumerate(data, 1):
                    ws[f'A{idx}'] = note
            else:
                # Handle tables
                if data:
                    row_idx = 1
                    for table in data:
                        for row in table:
                            for col_idx, cell_value in enumerate(row, 1):
                                if cell_value:
                                    ws.cell(row=row_idx, column=col_idx, value=str(cell_value))
                            row_idx += 1
                        row_idx += 2  # Add space between tables
                else:
                    ws['A1'] = f"No {sheet_name.lower()} data found"
        
        # Save to BytesIO for Streamlit download
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer
        
    except Exception as e:
        print(f"Error exporting to Excel: {e}")
        return None

def create_formatted_excel(extracted_data):
    """
    Create a formatted Excel file with styling.
    """
    try:
        wb = Workbook()
        wb.remove(wb.active)
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        sheets_data = {
            'Balance Sheet': extracted_data.get('balance_sheet', []),
            'Profit & Loss': extracted_data.get('profit_loss', []),
            'Cash Flow': extracted_data.get('cash_flow', []),
            'Notes': extracted_data.get('notes', [])
        }
        
        for sheet_name, data in sheets_data.items():
            ws = wb.create_sheet(title=sheet_name)
            
            # Add header
            ws['A1'] = sheet_name
            ws['A1'].font = header_font
            ws['A1'].fill = header_fill
            
            if sheet_name == 'Notes':
                for idx, note in enumerate(data, 3):
                    ws[f'A{idx}'] = note
            else:
                if data:
                    row_idx = 3
                    for table in data:
                        for row in table:
                            for col_idx, cell_value in enumerate(row, 1):
                                if cell_value:
                                    ws.cell(row=row_idx, column=col_idx, value=str(cell_value))
                            row_idx += 1
                        row_idx += 2
                else:
                    ws['A3'] = f"No {sheet_name.lower()} data found"
        
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer
        
    except Exception as e:
        print(f"Error creating formatted Excel: {e}")
        return None
